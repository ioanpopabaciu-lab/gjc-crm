"""
Import toate companiile din Excel ca Leads B2B in MongoDB GJC CRM
Fisier: data/Candidati cu viza plasati si programari .xlsx

Sheet 1: Angajator Initial (col 4), Angajator Final (col 5)
Sheet 2: Angajator Initial (col 5), Angajator Final (col 6)
"""

import asyncio
import os
from datetime import datetime
from motor.motor_asyncio import AsyncIOMotorClient
import openpyxl
from dotenv import load_dotenv

load_dotenv()
MONGO_URL = os.getenv("MONGO_URL", "")
DB_NAME = os.getenv("DB_NAME", "gjc_crm_db")

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "data", "Candidati cu viza plasati si programari .xlsx")


def normalize_company(name):
    if not name:
        return ""
    return name.strip().lower().replace("&", "and").replace("  ", " ")


async def main():
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws1 = wb[wb.sheetnames[0]]  # Muncitori ajunsi in Romania
    ws2 = wb[wb.sheetnames[1]]  # Muncitori programati

    rows1 = list(ws1.iter_rows(values_only=True))
    rows2 = list(ws2.iter_rows(values_only=True))

    # Colecteaza toate companiile unice din ambele sheet-uri
    companies_set = set()
    for r in rows1[1:]:
        if r[0]:
            if r[4] and str(r[4]).strip(): companies_set.add(str(r[4]).strip())  # Angajator Initial
            if r[5] and str(r[5]).strip(): companies_set.add(str(r[5]).strip())  # Angajator Final
    for r in rows2[1:]:
        if r[0]:
            if r[5] and str(r[5]).strip(): companies_set.add(str(r[5]).strip())  # Angajator Initial
            if r[6] and str(r[6]).strip(): companies_set.add(str(r[6]).strip())  # Angajator Final

    print(f"\n=== COMPANII GASITE IN EXCEL ({len(companies_set)}) ===")
    for c in sorted(companies_set):
        print(f"  - {c}")

    # Citeste leads existente
    existing_leads = await db.leads.find({}, {"_id": 0}).to_list(None)
    existing_names_normalized = {normalize_company(l.get("company_name", "")): l for l in existing_leads}

    print(f"\n=== LEADS EXISTENTE IN DB: {len(existing_leads)} ===")

    added = 0
    skipped = 0

    for comp_name in sorted(companies_set):
        comp_norm = normalize_company(comp_name)

        # Verifica daca exista deja (fuzzy match)
        found = False
        for existing_norm, existing_lead in existing_names_normalized.items():
            if (existing_norm == comp_norm or
                comp_norm in existing_norm or
                existing_norm in comp_norm):
                found = True
                print(f"  EXISTA DEJA: '{comp_name}' -> '{existing_lead.get('company_name', '')}'")
                break

        if not found:
            import uuid
            new_lead = {
                "id": str(uuid.uuid4()),
                "company_name": comp_name,
                "contact_person": None,
                "phone": None,
                "email": None,
                "city": None,
                "source": "Excel GJC",
                "responsible": None,
                "industry": "Constructii",
                "positions_needed": None,
                "estimated_value": None,
                "stage": "castigat",
                "notes": "Importat automat — companie cu lucratori nepalezi plasati/programati",
                "created_at": datetime.utcnow().isoformat(),
            }
            await db.leads.insert_one(new_lead)
            print(f"  ADAUGAT LEAD: '{comp_name}'")
            added += 1
        else:
            skipped += 1

    print(f"\n=== REZULTAT ===")
    print(f"  Leads adaugate: {added}")
    print(f"  Deja existente (sarite): {skipped}")
    print(f"  Total leads acum: {await db.leads.count_documents({})}")

    client.close()


if __name__ == "__main__":
    asyncio.run(main())
