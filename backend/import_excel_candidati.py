"""
Import candidati si programari din Excel in MongoDB GJC CRM
Fisier: data/Candidati cu viza plasati si programari .xlsx

Sheet 1: Muncitori ajunsi in Romania (plasati)
Sheet 2: Muncitori programati (programari viitoare)
"""

import asyncio
import os
from datetime import datetime
from motor.motor_asyncio import AsyncIOMotorClient
import openpyxl
from dotenv import load_dotenv

load_dotenv()
MONGO_URL = os.getenv("MONGO_URL", "")
DB_NAME = os.getenv("DB_NAME", "gjc_crm_db")

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "data", "Candidati cu viza plasati si programari .xlsx")

# Normalizare nume companie pentru matching
def normalize_company(name):
    if not name:
        return ""
    return name.strip().lower().replace("&", "and").replace("  ", " ")

# Normalizare nume candidat
def normalize_name(name):
    if not name:
        return ""
    return name.strip()

async def main():
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws1 = wb[wb.sheetnames[0]]  # Muncitori ajunsi in Romania
    ws2 = wb[wb.sheetnames[1]]  # Muncitori programati

    rows1 = list(ws1.iter_rows(values_only=True))
    rows2 = list(ws2.iter_rows(values_only=True))

    # Colecteaza toate companiile unice din ambele sheet-uri
    companies_set = set()
    for r in rows1[1:]:
        if r[0]:
            if r[4]: companies_set.add(r[4].strip())  # Angajator Initial
            if r[5]: companies_set.add(r[5].strip())  # Angajator Final
    for r in rows2[1:]:
        if r[0]:
            if r[5]: companies_set.add(r[5].strip())  # Angajator Initial
            if r[6]: companies_set.add(r[6].strip())  # Angajator Final

    print(f"\n=== COMPANII IDENTIFICATE ({len(companies_set)}) ===")
    for c in sorted(companies_set):
        print(f"  - {c}")

    # Cauta/creeaza companiile in MongoDB
    company_map = {}  # name -> {id, name}
    existing_companies = await db.companies.find({}).to_list(None)

    for comp_name in companies_set:
        # Cauta companie existenta (fuzzy match pe prima parte a numelui)
        found = None
        comp_lower = normalize_company(comp_name)
        for ec in existing_companies:
            ec_lower = normalize_company(ec.get("name", ""))
            if ec_lower == comp_lower or comp_lower in ec_lower or ec_lower in comp_lower:
                found = ec
                break

        if found:
            company_map[comp_name] = {"id": str(found["_id"]), "name": found["name"]}
            print(f"  GASIT: '{comp_name}' -> '{found['name']}'")
        else:
            # Creeaza compania noua
            new_comp = {
                "name": comp_name,
                "industry": "Constructii",
                "status": "activ",
                "country": "Romania",
                "created_at": datetime.utcnow().isoformat(),
                "notes": "Importat automat din Excel GJC",
            }
            result = await db.companies.insert_one(new_comp)
            company_map[comp_name] = {"id": str(result.inserted_id), "name": comp_name}
            print(f"  CREAT: '{comp_name}'")

    # === SHEET 1: Muncitori ajunsi in Romania (plasati) ===
    print(f"\n=== SHEET 1: {len(rows1)-1} muncitori ajunsi ===")
    sheet1_imported = 0
    sheet1_updated = 0

    for r in rows1[1:]:
        if not r[0]:
            continue

        name = normalize_name(r[0])
        status_excel = r[1] or ""
        angajator_final = normalize_name(r[5]) if r[5] else ""
        angajator_initial = normalize_name(r[4]) if r[4] else ""

        company_name = angajator_final or angajator_initial
        company_info = company_map.get(company_name, {}) if company_name else {}

        # Determina status CRM
        if "Obtinut A.M" in status_excel:
            crm_status = "plasat"
        elif "Obtinut P.S" in status_excel:
            crm_status = "plasat"
        elif "In Lucru" in status_excel:
            crm_status = "activ"
        elif "Depus (IGI)" in status_excel:
            crm_status = "în procesare"
        elif "Pregatit" in status_excel:
            crm_status = "în procesare"
        else:
            crm_status = "activ"

        # Cauta candidat existent
        existing = await db.candidates.find_one({
            "$or": [
                {"first_name": {"$regex": name.split()[0] if name else "", "$options": "i"}},
                {"last_name": {"$regex": name.split()[-1] if name else "", "$options": "i"}},
            ]
        })

        parts = name.split()
        first_name = parts[0] if parts else name
        last_name = " ".join(parts[1:]) if len(parts) > 1 else ""

        candidate_data = {
            "nationality": "Nepal",
            "birth_country": "Nepal",
            "status": crm_status,
            "company_name": company_info.get("name", company_name),
            "company_id": company_info.get("id", ""),
            "excel_status": status_excel,
            "angajator_initial": angajator_initial,
        }

        if existing:
            await db.candidates.update_one(
                {"_id": existing["_id"]},
                {"$set": candidate_data}
            )
            sheet1_updated += 1
        else:
            new_cand = {
                "first_name": first_name,
                "last_name": last_name,
                "nationality": "Nepal",
                "birth_country": "Nepal",
                "status": crm_status,
                "company_name": company_info.get("name", company_name),
                "company_id": company_info.get("id", ""),
                "excel_status": status_excel,
                "angajator_initial": angajator_initial,
                "created_at": datetime.utcnow().isoformat(),
                "notes": f"Importat din Excel. Status Excel: {status_excel}",
            }
            await db.candidates.insert_one(new_cand)
            sheet1_imported += 1

    print(f"  Importati: {sheet1_imported}, Actualizati: {sheet1_updated}")

    # === SHEET 2: Muncitori programati (programari viitoare) ===
    print(f"\n=== SHEET 2: {len(rows2)-1} muncitori programati ===")
    sheet2_imported = 0
    sheet2_updated = 0

    for r in rows2[1:]:
        if not r[0]:
            continue

        name = normalize_name(r[0])
        status_excel = r[1] or "Neinceput"
        location = r[2] or ""
        ora = str(r[3]) if r[3] else ""
        data_prog = r[4]
        angajator_initial = normalize_name(r[5]) if r[5] else ""
        angajator_final = normalize_name(r[6]) if r[6] else ""

        # Formateaza data
        if isinstance(data_prog, datetime):
            data_str = data_prog.strftime("%Y-%m-%d")
        elif data_prog:
            data_str = str(data_prog)[:10]
        else:
            data_str = ""

        # Formateaza ora
        if ":" in str(ora):
            ora_str = str(ora)[:5]
        else:
            ora_str = ""

        company_name = angajator_initial or angajator_final
        company_info = company_map.get(company_name, {}) if company_name else {}

        parts = name.split()
        first_name = parts[0] if parts else name
        last_name = " ".join(parts[1:]) if len(parts) > 1 else ""

        # Cauta candidat existent
        existing = await db.candidates.find_one({
            "$or": [
                {"first_name": {"$regex": f"^{parts[0]}" if parts else "", "$options": "i"}},
            ],
            "last_name": {"$regex": parts[-1] if len(parts) > 1 else parts[0] if parts else "", "$options": "i"}
        }) if parts else None

        candidate_data = {
            "nationality": "Nepal",
            "birth_country": "Nepal",
            "status": "în procesare",
            "company_name": company_info.get("name", company_name),
            "company_id": company_info.get("id", ""),
            "appointment_date": data_str,
            "appointment_time": ora_str,
            "appointment_location": location,
            "excel_status": status_excel,
        }

        if existing:
            await db.candidates.update_one(
                {"_id": existing["_id"]},
                {"$set": candidate_data}
            )
            sheet2_updated += 1
            cand_id = str(existing["_id"])
            cand_full_name = f"{existing.get('first_name', '')} {existing.get('last_name', '')}".strip()
        else:
            new_cand = {
                "first_name": first_name,
                "last_name": last_name,
                "nationality": "Nepal",
                "birth_country": "Nepal",
                "status": "în procesare",
                "company_name": company_info.get("name", company_name),
                "company_id": company_info.get("id", ""),
                "appointment_date": data_str,
                "appointment_time": ora_str,
                "appointment_location": location,
                "excel_status": status_excel,
                "created_at": datetime.utcnow().isoformat(),
                "notes": f"Programare IGI: {data_str} {ora_str} - {location}. Import Excel.",
            }
            result = await db.candidates.insert_one(new_cand)
            cand_id = str(result.inserted_id)
            cand_full_name = name
            sheet2_imported += 1

        # Creeaza/actualizeaza dosar imigrare cu data programarii
        existing_case = await db.immigration_cases.find_one({
            "candidate_name": {"$regex": parts[0] if parts else "", "$options": "i"},
            "company_name": {"$regex": angajator_initial[:5] if len(angajator_initial) > 5 else angajator_initial, "$options": "i"}
        }) if parts and angajator_initial else None

        case_data = {
            "candidate_name": cand_full_name or name,
            "candidate_id": cand_id,
            "company_name": company_info.get("name", company_name),
            "company_id": company_info.get("id", ""),
            "appointment_date": data_str,
            "appointment_time": ora_str,
            "appointment_location": location,
            "case_type": "Permis de muncă",
            "status": "activ",
            "nationality": "Nepal",
        }

        if existing_case:
            await db.immigration_cases.update_one(
                {"_id": existing_case["_id"]},
                {"$set": case_data}
            )
        else:
            case_data["current_stage"] = 1
            case_data["current_stage_name"] = "Recrutat"
            case_data["created_at"] = datetime.utcnow().isoformat()
            await db.immigration_cases.insert_one(case_data)

    print(f"  Importati: {sheet2_imported}, Actualizati: {sheet2_updated}")
    print("\n=== IMPORT COMPLET ===")
    client.close()

if __name__ == "__main__":
    asyncio.run(main())
